#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  compliance 0.py
#  
#  Copyright 2017 hazmanyusoff
#  

import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
import pandas as pd
import time
from datetime import datetime, date
import os.path
import collections
from tkinter.filedialog import askopenfilename
from tkinter import Tk

root = Tk()
root.withdraw()

filenew = askopenfilename()
if filenew is '':
	print('No file selected\n\nExiting...')
	time.sleep(2)
	quit()

#file_path = os.path.dirname(filenew)+'/Non Compliance Risk-comments.xlsx'
file_path = os.path.splitext(filenew)[0]+'-comments.xlsx'

# Cell format
redFill = PatternFill(start_color='FFC7CE',
                      end_color='FFC7CE',
                      fill_type='solid')

today = datetime.today()
ERROR = ['Undefined','Proposed']
action_comment = 'Evidence of action can be found at <insert where evidence can be found or if not done select action status Abandon and make a brief note as to why it was not done>'

wb = openpyxl.load_workbook(filenew)

i=0
for sheet in wb.worksheets:
	# Getting the first row and all column name
	# Read excel with no header to get accurate row number and fill nan with 'N/A' for easy tracking
	df = pd.read_excel(filenew,sheetname=i,header=None).fillna('N/A') #(filenew,sheetname=int)
	# iterate rows into tuples then into list
	cycle=list(df.itertuples(index=False))
	# empty list for our column name
	columnlist=[]

	firstrow=0
	while firstrow < len(cycle):
		for colu in cycle[firstrow]:
			columnlist.append(colu)
		counter=collections.Counter(columnlist) # Counting occurences in the columnlist
		if counter['N/A'] < 5:
			break
		firstrow+=1
		del columnlist[:]

	columnlist = [x for x in columnlist if x != 'N/A']
	COLUMN_NO = dict((keys,int) for keys in columnlist)

	i+=1

	for rowNum in range(firstrow+1, sheet.max_row + 1):		
		for columnNum in range(2, sheet.max_column + 1):
			# setting up variables for loop
			cell = sheet.cell(row=rowNum, column=columnNum)
			
			if cell.value in COLUMN_NO:
				COLUMN_NO[cell.value] = columnNum
	
			#elif (cell.value == '' or cell.value is None or cell.value in ERROR) and (columnNum != COLUMN_NO['Completion Date']):
			elif cell.value in ERROR:
				cell.fill = redFill
				
			elif cell.value == action_comment: #columnNum == COLUMN_NO['Mitigation Action Response Comments'] and 
				action_status = sheet.cell(row=rowNum, column=COLUMN_NO['Action Status'])
				if action_status.value == 'Complete' or action_status.value == 'Abandon':
					cell.fill = redFill
					
			if columnNum == COLUMN_NO['Action Status']:# cell.value == 'Active':
				#print(cell.value)
				# action_date = sheet.cell(row=rowNum, column=COLUMN_NO['Due Date'])
				if cell.value == 'Active':
					action_date = sheet.cell(row=rowNum, column=COLUMN_NO['Due Date'])
					if action_date.value is None or action_date.value == '' or action_date.value < today:
						action_date.fill = redFill
					# 	continue
					# else:
					# 	if action_date.value < today:
					# 		action_date.fill = redFill
				elif cell.value == 'Proposed' or cell.value == 'Complete':
					action_date = sheet.cell(row=rowNum, column=COLUMN_NO['Due Date'])
					if action_date.value == '' or action_date.value is None:
						action_date.fill = redFill
					# print(cell.value)

wb.save(file_path)
